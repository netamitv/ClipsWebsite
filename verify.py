import os
import re
import requests
import datetime
import logging
import time
import json
from urllib.parse import urlparse, parse_qs
from flask import redirect, url_for, session, flash, request, jsonify, render_template, abort
from dotenv import load_dotenv
from openpyxl import Workbook, load_workbook
from banned_ips import init_banned_ips, is_ip_banned, ban_ip, unban_ip, get_banned_ips
from werkzeug.security import check_password_hash

load_dotenv()

logger = logging.getLogger(__name__)

# Discord config
DISCORD_CLIENT_ID = os.getenv("DISCORD_CLIENT_ID", "")
DISCORD_CLIENT_SECRET = os.getenv("DISCORD_CLIENT_SECRET", "")
DISCORD_REDIRECT_URI = os.getenv("DISCORD_REDIRECT_URI", "")
DISCORD_API_ENDPOINT = "https://discord.com/api/v10"

# YouTube API (idk free api is enough)
YOUTUBE_API_KEY = os.getenv("YOUTUBE_API_KEY", "")
YOUTUBE_API_BASE_URL = "https://www.googleapis.com/youtube/v3"

# path to the excel
EXCEL_FILE = "verification_data.xlsx"


RATE_LIMIT_FILE = "discord_rate_limit.json"
MAX_REQUESTS_PER_HOUR = 100

# for verification
COOLDOWN_FILE = "verification_cooldowns.json"

# idiot cause its in the file not in an .env but you can recode it
ADMIN_USERNAME = "admin"
ADMIN_PASSWORD_HASH = "pbkdf2:sha256:260000$vGKtGO5E1pZgakVl$4f3c0a0373c3e9ecef95294ffd5d5641c9b6c0ec9c9302d9d17a2fdcc4432043"
def get_client_ip():
    """Get the client's IP address from the request."""
    if request.environ.get('HTTP_X_FORWARDED_FOR'):
        # For users behind a proxy
        ip = request.environ.get('HTTP_X_FORWARDED_FOR').split(',')[0]
    else:
        ip = request.environ.get('REMOTE_ADDR')
    return ip

def setup_verification(app):
    """Initialisiert die Verifizierungsfunktionen mit der Flask-App."""
    app.add_url_rule('/auth/discord/login', 'discord_login', discord_login)
    app.add_url_rule('/auth/discord/callback', 'discord_callback', discord_callback)
    app.add_url_rule('/auth/logout', 'auth_logout', auth_logout)
    app.add_url_rule('/verify-video', 'verify_video', verify_video, methods=['POST'])

    app.add_url_rule('/admin/ban-ip', 'admin_ban_ip', admin_ban_ip, methods=['POST'])
    app.add_url_rule('/admin/unban-ip', 'admin_unban_ip', admin_unban_ip, methods=['POST'])
    app.add_url_rule('/admin/banned-ips', 'admin_banned_ips', admin_banned_ips, methods=['GET', 'POST'])
    
    global logger
    logger = app.logger

    init_rate_limit_tracking()

    init_cooldown_tracking()

    init_banned_ips()

    @app.before_request
    def check_if_banned():
        if request.path.startswith('/admin'):
            return None
            
        ip = get_client_ip()
        ban = is_ip_banned(ip)
        if ban:
            return render_template('banned.html', reason=ban["reason"]), 403
    
    logger.info("Verification routes registered")

def init_rate_limit_tracking():
    """Initialisiert das Rate-Limit-Tracking."""
    if not os.path.exists(RATE_LIMIT_FILE):
        rate_limit_data = {
            "requests": [],
            "blocked_until": None
        }
        with open(RATE_LIMIT_FILE, 'w') as f:
            json.dump(rate_limit_data, f)

def init_cooldown_tracking():
    """Initialisiert das Cooldown-Tracking."""
    if not os.path.exists(COOLDOWN_FILE):
        with open(COOLDOWN_FILE, 'w') as f:
            json.dump({}, f)
        logger.info(f"Cooldown-Datei {COOLDOWN_FILE} erstellt")

def check_rate_limit():
    """Überprüft, ob wir das Rate-Limit erreicht haben."""
    try:
        with open(RATE_LIMIT_FILE, 'r') as f:
            rate_limit_data = json.load(f)
        
        if rate_limit_data.get("blocked_until"):
            blocked_until = datetime.datetime.fromisoformat(rate_limit_data["blocked_until"])
            if datetime.datetime.now() < blocked_until:
                remaining_time = blocked_until - datetime.datetime.now()
                minutes = int(remaining_time.total_seconds() // 60)
                return False, f"Discord API ist derzeit nicht verfügbar. Bitte versuche es in {minutes} Minuten erneut."
            else:
                rate_limit_data["blocked_until"] = None
        
        current_time = datetime.datetime.now()
        one_hour_ago = current_time - datetime.timedelta(hours=1)
        rate_limit_data["requests"] = [
            req for req in rate_limit_data["requests"] 
            if datetime.datetime.fromisoformat(req) > one_hour_ago
        ]
        
        if len(rate_limit_data["requests"]) >= MAX_REQUESTS_PER_HOUR:
            blocked_until = current_time + datetime.timedelta(minutes=30)
            rate_limit_data["blocked_until"] = blocked_until.isoformat()
            with open(RATE_LIMIT_FILE, 'w') as f:
                json.dump(rate_limit_data, f)
            return False, "Discord API-Limit erreicht. Bitte versuche es in 30 Minuten erneut."
        
        rate_limit_data["requests"].append(current_time.isoformat())
        with open(RATE_LIMIT_FILE, 'w') as f:
            json.dump(rate_limit_data, f)
        
        return True, None
    except Exception as e:
        logger.error(f"Fehler beim Überprüfen des Rate-Limits: {str(e)}", exc_info=True)
        return True, None

def update_rate_limit_block(retry_after):
    """Aktualisiert die Rate-Limit-Blockierung basierend auf der Discord-Antwort."""
    try:
        with open(RATE_LIMIT_FILE, 'r') as f:
            rate_limit_data = json.load(f)
        
        blocked_until = datetime.datetime.now() + datetime.timedelta(seconds=retry_after)
        rate_limit_data["blocked_until"] = blocked_until.isoformat()
        
        with open(RATE_LIMIT_FILE, 'w') as f:
            json.dump(rate_limit_data, f)
        
        logger.warning(f"Rate-Limit-Blockierung gesetzt bis: {blocked_until}")
    except Exception as e:
        logger.error(f"Fehler beim Aktualisieren der Rate-Limit-Blockierung: {str(e)}", exc_info=True)

def discord_login():
    """Route für die Discord-Anmeldung."""
    can_proceed, error_message = check_rate_limit()
    if not can_proceed:
        flash(error_message, "error")
        return redirect(url_for('verification'))
    
    # 0Auth2 Url creation with Scopes
    scope = "identify email connections"
    discord_login_url = f"{DISCORD_API_ENDPOINT}/oauth2/authorize?client_id={DISCORD_CLIENT_ID}&redirect_uri={DISCORD_REDIRECT_URI}&response_type=code&scope={scope}"
    logger.info(f"Redirecting to Discord login URL")
    return redirect(discord_login_url)

def discord_callback():
    """Callback-Route für die Discord-Anmeldung."""
    code = request.args.get('code')
    if not code:
        logger.error("No code received from Discord")
        flash("Kein Code von Discord erhalten", "error")
        return redirect(url_for('verification'))
    
    logger.info(f"Received code from Discord")
    
    can_proceed, error_message = check_rate_limit()
    if not can_proceed:
        flash(error_message, "error")
        return redirect(url_for('verification'))
    
    data = {
        'client_id': DISCORD_CLIENT_ID,
        'client_secret': DISCORD_CLIENT_SECRET,
        'grant_type': 'authorization_code',
        'code': code,
        'redirect_uri': DISCORD_REDIRECT_URI,
        'scope': 'identify email connections'
    }
    headers = {
        'Content-Type': 'application/x-www-form-urlencoded'
    }
    
    try:
        logger.info("Requesting token from Discord")
        response = requests.post(f"{DISCORD_API_ENDPOINT}/oauth2/token", data=data, headers=headers)
        logger.info(f"Token response status: {response.status_code}")
        
        if response.status_code == 429:
            retry_after = int(response.headers.get('Retry-After', 60))
            logger.warning(f"Discord API rate limit reached. Retry after: {retry_after} seconds")
            update_rate_limit_block(retry_after)
            
            flash("Discord API-Limit erreicht. Bitte versuche es später erneut.", "error")
            return redirect(url_for('verification'))
        
        response.raise_for_status()
        
        tokens = response.json()
        access_token = tokens.get('access_token')
        
        logger.info("Requesting user information from Discord")
        user_response = requests.get(f"{DISCORD_API_ENDPOINT}/users/@me", 
                                    headers={'Authorization': f"Bearer {access_token}"})
        logger.info(f"User info response status: {user_response.status_code}")
        
        if user_response.status_code == 429:
            retry_after = int(user_response.headers.get('Retry-After', 60))
            logger.warning(f"Discord API rate limit reached. Retry after: {retry_after} seconds")
            update_rate_limit_block(retry_after)
            
            flash("Discord API-Limit erreicht. Bitte versuche es später erneut.", "error")
            return redirect(url_for('verification'))
        
        user_response.raise_for_status()
        
        user_data = user_response.json()
        
        logger.info("Requesting user connections from Discord")
        connections_response = requests.get(f"{DISCORD_API_ENDPOINT}/users/@me/connections", 
                                          headers={'Authorization': f"Bearer {access_token}"})
        logger.info(f"Connections response status: {connections_response.status_code}")
        
        if connections_response.status_code == 429:
            retry_after = int(connections_response.headers.get('Retry-After', 60))
            logger.warning(f"Discord API rate limit reached. Retry after: {retry_after} seconds")
            update_rate_limit_block(retry_after)
            
            flash("Discord API-Limit erreicht. Bitte versuche es später erneut.", "error")
            return redirect(url_for('verification'))
        
        connections_response.raise_for_status()
        
        connections = connections_response.json()
        
        session['user'] = user_data
        session['connections'] = connections
        
        session.permanent = True
        
        if connections:
            connection_names = [f"{conn['type'].capitalize()} ({conn['name']})" for conn in connections]
            accounts_text = ", ".join(connection_names)
            flash(f"Erfolgreich mit Discord angemeldet! Du hast folgende Accounts verbunden: {accounts_text}", "success")
        else:
            flash("Erfolgreich mit Discord angemeldet! Du hast keine Accounts mit deinem Discord-Profil verbunden.", "success")
        
        logger.info(f"User {user_data['username']} successfully logged in with Discord")
        
    except requests.exceptions.HTTPError as e:
        logger.error(f"HTTP error during Discord login: {str(e)}", exc_info=True)
        flash(f"Fehler bei der Discord-Anmeldung: {str(e)}", "error")
        return redirect(url_for('verification'))
    except Exception as e:
        logger.error(f"Error during Discord login: {str(e)}", exc_info=True)
        flash(f"Fehler bei der Discord-Anmeldung: {str(e)}", "error")
        return redirect(url_for('verification'))
    
    return redirect(url_for('dashboard'))

def verify_video():
    """Route für die Verifizierung eines Videos."""
    if 'user' not in session:
        flash("Du musst angemeldet sein, um ein Video zu verifizieren.", "error")
        return redirect(url_for('verification'))
    
    platform = request.form.get('platform')
    video_link = request.form.get('video_link')
    
    if not platform or not video_link:
        flash("Bitte fülle alle Felder aus.", "error")
        return redirect(url_for('dashboard'))
    
    user_id = session['user']['id']
    
    cooldowns = {}
    if os.path.exists(COOLDOWN_FILE):
        try:
            with open(COOLDOWN_FILE, 'r') as f:
                cooldowns = json.load(f)
        except Exception as e:
            logger.error(f"Fehler beim Laden der Cooldown-Datei: {str(e)}", exc_info=True)
    
    cooldown_key = f"{user_id}_{platform}"
    current_time = datetime.datetime.now()
    
    if cooldown_key in cooldowns:
        last_attempt_time = datetime.datetime.fromisoformat(cooldowns[cooldown_key])
        time_diff = current_time - last_attempt_time
        
        if time_diff.total_seconds() < 24 * 60 * 60:
            remaining_seconds = 24 * 60 * 60 - time_diff.total_seconds()
            remaining_hours = int(remaining_seconds // 3600)
            remaining_minutes = int((remaining_seconds % 3600) // 60)
            
            session[f"last_attempt_{platform}"] = cooldowns[cooldown_key]
            
            session['verification_status'] = {
                'success': False,
                'message': f"Du kannst nur alle 24 Stunden einen Verifizierungsversuch für {platform.capitalize()} durchführen.",
                'cooldown': {
                    'platform': platform,
                    'remaining_hours': remaining_hours,
                    'remaining_minutes': remaining_minutes,
                    'expires_at': (current_time + datetime.timedelta(seconds=remaining_seconds)).isoformat()
                }
            }
            return redirect(url_for('dashboard'))
    
    cooldowns[cooldown_key] = current_time.isoformat()
    
    session[f"last_attempt_{platform}"] = current_time.isoformat()
    
    try:
        with open(COOLDOWN_FILE, 'w') as f:
            json.dump(cooldowns, f)
    except Exception as e:
        logger.error(f"Fehler beim Speichern der Cooldown-Datei: {str(e)}", exc_info=True)
    
    video_info = extract_video_info(platform, video_link)
    
    if not video_info:
        session['verification_status'] = {
            'success': False,
            'message': f"Ungültiger {platform.capitalize()}-Link. Bitte überprüfe den Link und versuche es erneut."
        }
        return redirect(url_for('dashboard'))
    
    user_data = session['user']
    connections = session['connections']
    
    if platform == 'tiktok':
        platform_username = video_info.get('username')
        if not platform_username:
            session['verification_status'] = {
                'success': False,
                'message': "Konnte den TikTok-Benutzernamen nicht extrahieren."
            }
            return redirect(url_for('dashboard'))
        
        tiktok_connection = next((conn for conn in connections if conn['type'].lower() == 'tiktok'), None)
        
        if not tiktok_connection:
            session['verification_status'] = {
                'success': False,
                'message': "Du hast keinen TikTok-Account mit deinem Discord-Profil verbunden."
            }
            return redirect(url_for('dashboard'))
        
        if tiktok_connection['name'].lower() != platform_username.lower():
            session['verification_status'] = {
                'success': False,
                'message': f"Der TikTok-Account im Video (@{platform_username}) stimmt nicht mit deinem verknüpften TikTok-Account (@{tiktok_connection['name']}) überein."
            }
            return redirect(url_for('dashboard'))
        
        session['verification_status'] = {
            'success': True,
            'message': f"Dein TikTok-Account @{platform_username} wurde erfolgreich verifiziert!"
        }
        
        send_webhook_notification(user_data, platform, platform_username, video_link)
        
    elif platform == 'youtube':
        video_id = video_info.get('video_id')
        if not video_id:
            session['verification_status'] = {
                'success': False,
                'message': "Konnte die YouTube-Video-ID nicht extrahieren."
            }
            return redirect(url_for('dashboard'))
        
        youtube_connection = next((conn for conn in connections if conn['type'].lower() == 'youtube'), None)
        
        if not youtube_connection:
            session['verification_status'] = {
                'success': False,
                'message': "Du hast keinen YouTube-Account mit deinem Discord-Profil verbunden."
            }
            return redirect(url_for('dashboard'))
        
        try:
            if YOUTUBE_API_KEY:
                video_url = f"{YOUTUBE_API_BASE_URL}/videos?part=snippet&id={video_id}&key={YOUTUBE_API_KEY}"
                response = requests.get(video_url)
                response.raise_for_status()
                
                video_data = response.json()
                if video_data.get('items'):
                    snippet = video_data['items'][0]['snippet']
                    channel_title = snippet['channelTitle']
                    
                    if youtube_connection['name'].lower() != channel_title.lower():
                        session['verification_status'] = {
                            'success': False,
                            'message': f"Der YouTube-Kanal des Videos ({channel_title}) stimmt nicht mit deinem verknüpften YouTube-Account ({youtube_connection['name']}) überein."
                        }
                        return redirect(url_for('dashboard'))
                    
                    session['verification_status'] = {
                        'success': True,
                        'message': f"Dein YouTube-Kanal '{channel_title}' wurde erfolgreich verifiziert!"
                    }
                    
                    send_webhook_notification(user_data, platform, channel_title, video_link)
                else:
                    session['verification_status'] = {
                        'success': False,
                        'message': "Das YouTube-Video wurde nicht gefunden."
                    }
                    return redirect(url_for('dashboard'))
            else:
                session['verification_status'] = {
                    'success': False,
                    'message': "YouTube API-Schlüssel nicht konfiguriert. Bitte kontaktiere den Administrator."
                }
                return redirect(url_for('dashboard'))
        except Exception as e:
            logger.error(f"Fehler bei der Überprüfung des YouTube-Videos: {str(e)}", exc_info=True)
            session['verification_status'] = {
                'success': False,
                'message': f"Fehler bei der Überprüfung des YouTube-Videos: {str(e)}"
            }
            return redirect(url_for('dashboard'))
    
    else:
        session['verification_status'] = {
            'success': False,
            'message': "Plattform nicht unterstützt."
        }
    
    return redirect(url_for('dashboard'))

def extract_video_info(platform, video_link):
    """Extrahiert Informationen aus einem Video-Link."""
    if platform == 'tiktok':
        # TikTok-Link-Format: https://www.tiktok.com/@username/video/1234567890123456789
        tiktok_pattern = r'tiktok\.com/@([^/]+)/video/(\d+)'
        match = re.search(tiktok_pattern, video_link)
        
        if match:
            return {
                'username': match.group(1),
                'video_id': match.group(2)
            }
        
        # TikTok-shortlink-Format: https://vm.tiktok.com/XXXXXXXX/
        if 'vm.tiktok.com' in video_link:
            try:
                # Folge dem Redirect, um den tatsächlichen Link zu erhalten
                response = requests.head(video_link, allow_redirects=True)
                final_url = response.url
                
                match = re.search(tiktok_pattern, final_url)
                if match:
                    return {
                        'username': match.group(1),
                        'video_id': match.group(2)
                    }
            except Exception as e:
                logger.error(f"Fehler beim Folgen des TikTok-Kurzlinks: {str(e)}", exc_info=True)
    
    elif platform == 'youtube':
        # YouTube-Link-Format: https://www.youtube.com/watch?v=xxxxxxxxxxx
        parsed_url = urlparse(video_link)
        
        if 'youtube.com' in parsed_url.netloc and parsed_url.path == '/watch':
            query_params = parse_qs(parsed_url.query)
            video_id = query_params.get('v', [None])[0]
            
            if video_id:
                return {
                    'video_id': video_id
                }
        
        # YouTube-Shorts-Format: https://www.youtube.com/shorts/xxxxxxxxxxx
        elif 'youtube.com' in parsed_url.netloc and '/shorts/' in parsed_url.path:
            video_id = parsed_url.path.split('/shorts/')[1]
            
            if video_id:
                return {
                    'video_id': video_id
                }
        
        # YouTube-Shortlink-Format: https://youtu.be/xxxxxxxxxxx
        elif 'youtu.be' in parsed_url.netloc:
            video_id = parsed_url.path.lstrip('/')
            
            if video_id:
                return {
                    'video_id': video_id
                }
    
    return None

def send_webhook_notification(user_data, platform, platform_username, video_link):
    """Speichert die Verifizierungsdaten in einer Excel-Tabelle anstatt an einen Webhook zu senden."""
    discord_id = user_data['id']
    discord_username = user_data['username']
    
    user_ip = get_client_ip()
    
    verification_time = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    
    if os.path.exists(EXCEL_FILE):
        try:
            wb = load_workbook(EXCEL_FILE)
            ws = wb.active
        except Exception as e:
            logger.error(f"Fehler beim Laden der Excel-Datei: {str(e)}", exc_info=True)
            wb = Workbook()
            ws = wb.active
            ws.title = "Verifizierungen"
            ws.append(["Datum", "Discord ID", "Discord Username", "Plattform", "Plattform Username", "Video Link", "IP Adresse"])
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Verifizierungen"
        ws.append(["Datum", "Discord ID", "Discord Username", "Plattform", "Plattform Username", "Video Link", "IP Adresse"])
    
    ws.append([verification_time, discord_id, discord_username, platform, platform_username, video_link, user_ip])
    
    try:
        # Speichere die Datei
        wb.save(EXCEL_FILE)
        logger.info(f"Verifizierungsdaten für {discord_username} in Excel-Tabelle gespeichert")
    except Exception as e:
        logger.error(f"Fehler beim Speichern der Excel-Datei: {str(e)}", exc_info=True)

def auth_logout():
    """Route für die Abmeldung."""
    session.pop('user', None)
    session.pop('connections', None)
    
    for key in list(session.keys()):
        if key.startswith('last_attempt_'):
            session.pop(key, None)
    
    flash("Du wurdest erfolgreich abgemeldet.", "success")
    return redirect(url_for('verification'))

# Admin routes for managing banned IPs
def admin_banned_ips():
    """Admin page to view and manage banned IPs."""
    error = None
    authenticated = False
    banned_ips = []
    if 'admin_authenticated' in session and session['admin_authenticated']:
        authenticated = True
        banned_data = get_banned_ips()
        banned_ips = banned_data["banned_ips"]
    
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        
        if username == ADMIN_USERNAME and check_password_hash(ADMIN_PASSWORD_HASH, password):
            session['admin_authenticated'] = True
            authenticated = True
            banned_data = get_banned_ips()
            banned_ips = banned_data["banned_ips"]
        else:
            error = "Ungültiger Benutzername oder Passwort"
    
    return render_template('admin/banned_ips.html', 
                          authenticated=authenticated, 
                          error=error,
                          banned_ips=banned_ips)

def admin_ban_ip():
    """Admin route to ban an IP."""
    if not session.get('admin_authenticated'):
        flash("Du musst angemeldet sein, um diese Aktion durchzuführen.", "error")
        return redirect(url_for('admin_banned_ips'))
    
    ip_address = request.form.get('ip_address')
    reason = request.form.get('reason', 'No reason provided')
    duration_days = request.form.get('duration')
    
    if duration_days:
        try:
            duration_days = int(duration_days)
        except ValueError:
            duration_days = None
    
    if not ip_address:
        flash("IP address is required", "error")
        return redirect(url_for('admin_banned_ips'))
    
    if ban_ip(ip_address, reason, duration_days):
        flash(f"IP {ip_address} has been banned successfully", "success")
    else:
        flash(f"Error banning IP {ip_address}", "error")
    
    return redirect(url_for('admin_banned_ips'))

def admin_unban_ip():
    """Admin route to unban an IP."""
    if not session.get('admin_authenticated'):
        flash("Du musst angemeldet sein, um diese Aktion durchzuführen.", "error")
        return redirect(url_for('admin_banned_ips'))
    
    ip_address = request.form.get('ip_address')
    
    if not ip_address:
        flash("IP address is required", "error")
        return redirect(url_for('admin_banned_ips'))
    
    if unban_ip(ip_address):
        flash(f"IP {ip_address} has been unbanned successfully", "success")
    else:
        flash(f"Error unbanning IP {ip_address}", "error")
    
    return redirect(url_for('admin_banned_ips'))
