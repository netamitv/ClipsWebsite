import os
import logging
import datetime
import json
from logging.handlers import RotatingFileHandler
from flask import Flask, render_template, send_from_directory, request, redirect, url_for, session, flash, jsonify, send_file
from dotenv import load_dotenv
from verify import setup_verification, RATE_LIMIT_FILE
from werkzeug.security import check_password_hash, generate_password_hash
import openpyxl

load_dotenv()

app = Flask(__name__)
app.config['SECRET_KEY'] = os.getenv('SECRET_KEY', 'your-secure-secret-key-here')
app.config['PERMANENT_SESSION_LIFETIME'] = datetime.timedelta(days=30)

# Maintenance Mode 
MAINTENANCE_MODE = False  # True = Maintenance on, False = Maintenance off

@app.before_request
def check_maintenance():
    if MAINTENANCE_MODE and request.endpoint not in ['static', 'favicon']:
        return render_template('maintenance.html'), 503

setup_verification(app)

# Favicon
@app.route('/favicon.ico')
def favicon():
    return send_from_directory(os.path.join(app.root_path, 'static'),
                              'favicon.ico', mimetype='image/vnd.microsoft.icon')

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/about')
@app.route('/about/')
def about():
    return render_template('about.html')

@app.route('/contact')
@app.route('/contact/')
def contact():
    return render_template('contact.html')

@app.route('/terms')
@app.route('/terms/')
def terms():
    return render_template('terms.html')

@app.route('/privacy')
@app.route('/privacy/')
def privacy():
    return render_template('privacy.html')

@app.route('/tutorial')
@app.route('/tutorial/')
def tutorial():
    return render_template('header/tutorial.html')

@app.route('/tutorial/phone.html')
def tutorial_phone():
    return render_template('tutorial/phone.html')

@app.route('/tutorial/pc.html')
def tutorial_pc():
    return render_template('tutorial/pc.html')

@app.route('/tutorial/video-apps')
@app.route('/tutorial/video-apps.html')
def tutorial_video_apps():
    return render_template('tutorial/video-apps.html')

@app.route('/upload')
@app.route('/upload/')
def upload():
    return render_template('header/upload.html')

@app.route('/format-ideas')
@app.route('/format-ideas/')
def format_ideas():
    return render_template('header/format-ideas.html')

@app.route('/monetization')
@app.route('/monetization/')
def monetization():
    return render_template('header/monetization.html')

@app.route('/dashboard')
@app.route('/dashboard/')
def dashboard():
    # checks if the user is logged in
    if 'user' not in session:
        flash("Du musst angemeldet sein, um das Dashboard zu verwenden.", "error")
        return redirect(url_for('verification'))
    
    user = session['user']
    connections = session.get('connections', [])
    verification_status = session.pop('verification_status', None)
    
    verified_count = sum(1 for conn in connections if conn.get('verified', False))
    
    # load data from the excel
    total_verifications = 0
    recent_activities = []
    
    excel_file = os.path.join(app.root_path, 'verification_data.xlsx')
    if os.path.exists(excel_file):
        try:
            wb = openpyxl.load_workbook(excel_file, read_only=True)
            ws = wb.active
            
            user_verifications = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[1] == user['id']:
                    user_verifications.append({
                        'date': row[0],
                        'platform': row[3],
                        'username': row[4],
                        'video_link': row[5]
                    })
            
            total_verifications = len(user_verifications)
            
            for verification in user_verifications[-5:]:
                recent_activities.append({
                    'type': 'verification',
                    'title': f'{verification["platform"].title()} Account verifiziert',
                    'description': f'@{verification["username"]} wurde erfolgreich verifiziert',
                    'time': verification['date'].strftime('%d.%m.%Y %H:%M') if hasattr(verification['date'], 'strftime') else str(verification['date'])
                })
                
        except Exception as e:
            app.logger.error(f"Fehler beim Laden der Excel-Daten: {str(e)}")
    
    discord_id = int(user['id'])
    discord_epoch = 1420070400000
    timestamp = ((discord_id >> 22) + discord_epoch) / 1000
    creation_date = datetime.datetime.fromtimestamp(timestamp)
    days_since_join = (datetime.datetime.now() - creation_date).days
    
    # load aktive cooldown
    cooldowns = {}
    user_id = user['id']
    current_time = datetime.datetime.now()
    
    cooldown_file = os.path.join(app.root_path, 'verification_cooldowns.json')
    if os.path.exists(cooldown_file):
        try:
            with open(cooldown_file, 'r') as f:
                all_cooldowns = json.load(f)
                
                for key, timestamp in all_cooldowns.items():
                    if key.startswith(f"{user_id}_"):
                        platform = key.split('_')[1]
                        last_attempt_time = datetime.datetime.fromisoformat(timestamp)
                        time_diff = current_time - last_attempt_time
                        
                        if time_diff.total_seconds() < 24 * 60 * 60:
                            remaining_seconds = 24 * 60 * 60 - time_diff.total_seconds()
                            remaining_hours = int(remaining_seconds // 3600)
                            remaining_minutes = int((remaining_seconds % 3600) // 60)
                            
                            cooldowns[platform] = {
                                'remaining_hours': remaining_hours,
                                'remaining_minutes': remaining_minutes,
                                'expires_at': (current_time + datetime.timedelta(seconds=remaining_seconds)).isoformat()
                            }
        except Exception as e:
            app.logger.error(f"Fehler beim Laden der Cooldown-Datei: {str(e)}")
    
    if not recent_activities:
        recent_activities.append({
            'type': 'login',
            'title': 'Mit Discord angemeldet',
            'description': 'Erfolgreich mit Discord authentifiziert',
            'time': 'Heute'
        })
    
    # change session to permanent
    session.permanent = True
    
    return render_template('dashboard.html',
                          user=user,
                          connections=connections,
                          verification_status=verification_status,
                          cooldowns=cooldowns,
                          verified_count=verified_count,
                          total_verifications=total_verifications,
                          days_since_join=days_since_join,
                          recent_activities=recent_activities)

@app.route('/verification')
@app.route('/verification/')
def verification():
    logged_in = 'user' in session
    
    if logged_in:
        return redirect(url_for('dashboard'))
    
    app.logger.info(f"Verification Route - Logged in: {logged_in}")
    
    verification_status = session.pop('verification_status', None)
    
    discord_blocked_until = None
    if os.path.exists(RATE_LIMIT_FILE):
        try:
            with open(RATE_LIMIT_FILE, 'r') as f:
                rate_limit_data = json.load(f)
                if rate_limit_data.get("blocked_until"):
                    blocked_until = datetime.datetime.fromisoformat(rate_limit_data["blocked_until"])
                    if datetime.datetime.now() < blocked_until:
                        remaining_time = blocked_until - datetime.datetime.now()
                        minutes = int(remaining_time.total_seconds() // 60)
                        seconds = int(remaining_time.total_seconds() % 60)
                        discord_blocked_until = {
                            'time': blocked_until.strftime('%H:%M:%S'),
                            'minutes': minutes,
                            'seconds': seconds
                        }
        except Exception as e:
            app.logger.error(f"Fehler beim Lesen der Rate-Limit-Datei: {str(e)}")
    
    return render_template('header/verification.html', 
                          logged_in=logged_in, 
                          user=session.get('user'),
                          verification_status=verification_status,
                          cooldowns={},
                          discord_blocked_until=discord_blocked_until)

# special routs for error pages for direct display (idk im not a native englisch speaker im sry)
@app.route('/404-demo')
@app.route('/404-demo/')
def show_404_demo():
    return render_template('404.html')

@app.route('/500-demo')
@app.route('/500-demo/')
def show_500_demo():
    return render_template('500.html')

@app.errorhandler(404)
def page_not_found(e):
    return render_template('404.html'), 404

@app.errorhandler(500)
def server_error(e):
    return render_template('500.html'), 500

@app.after_request
def add_security_headers(response):
    response.headers['X-Content-Type-Options'] = 'nosniff'
    response.headers['X-Frame-Options'] = 'SAMEORIGIN'
    response.headers['X-XSS-Protection'] = '1; mode=block'
    return response

@app.route('/admin/logout', methods=['GET', 'POST'])
def admin_logout():
    session.pop('admin_authenticated', None)
    flash('Du wurdest erfolgreich abgemeldet.', 'success')
    return redirect(url_for('admin_excel'))

@app.route('/reset-rate-limit', methods=['POST'])
def reset_rate_limit():
    if not app.debug:
        return jsonify({"error": "Diese Funktion ist nur im Debug-Modus verfügbar"}), 403
    
    if os.path.exists(RATE_LIMIT_FILE):
        try:
            with open(RATE_LIMIT_FILE, 'r') as f:
                rate_limit_data = json.load(f)
            
            rate_limit_data["requests"] = []
            rate_limit_data["blocked_until"] = None
            
            with open(RATE_LIMIT_FILE, 'w') as f:
                json.dump(rate_limit_data, f)
            
            return jsonify({"success": True, "message": "Rate-Limit zurückgesetzt"})
        except Exception as e:
            return jsonify({"error": str(e)}), 500
    else:
        return jsonify({"error": "Rate-Limit-Datei nicht gefunden"}), 404

# admin login username
ADMIN_USERNAME = "admin"
ADMIN_PASSWORD_HASH = generate_password_hash("GHs1402aB!")  # change this to your password you wanna use

@app.route('/admin/excel', methods=['GET', 'POST'])
def admin_excel():
    """Admin-Seite zum Anzeigen der Excel-Datei mit Verifizierungsdaten."""
    error = None
    authenticated = False
    headers = []
    data = []
    t
    if 'admin_authenticated' in session and session['admin_authenticated']:
        authenticated = True
    
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        
        if username == ADMIN_USERNAME and check_password_hash(ADMIN_PASSWORD_HASH, password):
            session['admin_authenticated'] = True
            authenticated = True
        else:
            error = "Ungültiger Benutzername oder Passwort"
    
    if authenticated:
        excel_file = os.path.join(app.root_path, 'verification_data.xlsx')
        
        if os.path.exists(excel_file):
            try:
                wb = openpyxl.load_workbook(excel_file, read_only=True)
                ws = wb.active
                
                rows = list(ws.rows)
                if rows:
                    headers = [cell.value for cell in rows[0]]
                    data = [[cell.value for cell in row] for row in rows[1:]]
            except Exception as e:
                app.logger.error(f"Fehler beim Laden der Excel-Datei: {str(e)}")
                error = f"Fehler beim Laden der Excel-Datei: {str(e)}"
    
    return render_template('admin/excel.html', 
                          authenticated=authenticated, 
                          error=error,
                          headers=headers,
                          data=data)

@app.route('/admin/excel/download')
def admin_excel_download():
    """Route zum Herunterladen der Excel-Datei."""
    if not session.get('admin_authenticated'):
        return redirect(url_for('admin_excel'))
    
    excel_file = os.path.join(app.root_path, 'verification_data.xlsx')
    
    if not os.path.exists(excel_file):
        flash("Die Excel-Datei existiert nicht.", "error")
        return redirect(url_for('admin_excel'))
    
    return send_file(excel_file, 
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    as_attachment=True,
                    download_name='verification_data.xlsx')

if not app.debug:
    os.makedirs('logs', exist_ok=True)
    
    file_handler = RotatingFileHandler('logs/app.log', maxBytes=10240, backupCount=10)
    file_handler.setFormatter(logging.Formatter(
        '%(asctime)s %(levelname)s: %(message)s [in %(pathname)s:%(lineno)d]'
    ))
    file_handler.setLevel(logging.INFO)
    app.logger.addHandler(file_handler)
    app.logger.setLevel(logging.INFO)
    app.logger.info('Application startup')

application = app

if __name__ == '__main__':
    print("ClipsWebsite Server starting")
    print("HTTP Server: ")
    print("HTTP-Modus: No SSL-Warning")
    
    # easy http for https use a domain under cloudflare or smth idk
    app.run(host='0.0.0.0', port=10938, debug=False)
